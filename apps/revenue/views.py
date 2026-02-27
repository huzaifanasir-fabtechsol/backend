from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction, IntegrityError
from django.http import HttpResponse
from django.db.models import Sum, Q
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from apps.revenue.models import Car, CarCategory, Order, OrderItem, Customer, Saler, CompanyAccount, Auction, Transaction
from apps.revenue.serializers import CarSerializer, CarCategorySerializer, OrderSerializer, OrderItemSerializer, CreateOrderSerializer, CustomerSerializer, SalerSerializer, CompanyAccountSerializer, AuctionSerializer, TransactionSerializer
from project.pagination import CustomPageNumberPagination
from apps.expense.models import Expense
from django.conf import settings
from apps.account.models import User
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import Image
from reportlab.lib.utils import ImageReader

class CarCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CarCategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = CarCategory.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(company__icontains=search) |
                Q(description__icontains=search)
            )
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    @action(detail=False, methods=['get'])
    def all(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

class CarViewSet(viewsets.ModelViewSet):
    serializer_class = CarSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        return Car.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Order.objects.filter(user=self.request.user).order_by('-created_at')
        
        payment_status = self.request.query_params.get('payment_status')
        transaction_type = self.request.query_params.get('transaction_type')
        transaction_catagory = self.request.query_params.get('transaction_catagory')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        search = self.request.query_params.get('search')
        
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        if transaction_catagory:
            queryset = queryset.filter(transaction_catagory=transaction_catagory)
        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction_date__lte=end_date)
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        orders = Order.objects.filter(user=request.user)
        expenses = Expense.objects.filter(user=request.user)
        
        approved_amount = orders.filter(payment_status='completed').aggregate(total=Sum('total_amount'))['total'] or 0
        pending_amount = orders.filter(payment_status='pending').aggregate(total=Sum('total_amount'))['total'] or 0
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0
        total_purchase = orders.filter(transaction_type='purchase').aggregate(total=Sum('total_amount'))['total'] or 0
        latest_orders = orders.order_by('-transaction_date')[:10].values('transaction_date', 'transaction_type', 'payment_status', 'total_amount')
        
        return Response({
            'approved_amount': float(approved_amount),
            'pending_amount': float(pending_amount),
            'total_expense': float(total_expense),
            'total_purchase': float(total_purchase),
            'latest_orders': list(latest_orders)
        })

    def _calculate_item_total(self, item):
        order_type = self.request.data.get('transaction_type', 'sale')
        
        if order_type == 'nagare':
            return (
                item.get('vehicle_price', 0) +
                item.get('vehicle_price_tax', 0) +
                item.get('recycle_fee', 0) +
                item.get('listing_fee', 0) +
                item.get('listing_fee_tax', 0) +
                item.get('canceling_fee', 0) -
                item.get('successful_bid', 0) -
                item.get('successful_bid_tax', 0) -
                item.get('commission_fee', 0) -
                item.get('commission_fee_tax', 0) -
                item.get('transport_fee', 0) -
                item.get('transport_fee_tax', 0) -
                item.get('registration_fee', 0) -
                item.get('registration_fee_tax', 0)
            )
        else:
            return (
                item.get('vehicle_price', 0) +
                item.get('vehicle_price_tax', 0) +
                item.get('recycle_fee', 0) +
                item.get('canceling_fee', 0) -
                item.get('listing_fee', 0) -
                item.get('listing_fee_tax', 0) -
                item.get('successful_bid', 0) -
                item.get('successful_bid_tax', 0) -
                item.get('commission_fee', 0) -
                item.get('commission_fee_tax', 0) -
                item.get('transport_fee', 0) -
                item.get('transport_fee_tax', 0) -
                item.get('registration_fee', 0) -
                item.get('registration_fee_tax', 0)
            )

    def _build_order_item_payload(self, item_data):
        def to_decimal(value):
            if isinstance(value, str):
                value = value.replace(',', '')
            return float(value) if value else 0
        
        return {
            'venue': item_data.get('venue', ''),
            'notes': item_data.get('notes', ''),
            'vehicle_price': to_decimal(item_data.get('vehicle_price', 0)),
            'vehicle_price_tax': to_decimal(item_data.get('vehicle_price_tax', 0)),
            'recycle_fee': to_decimal(item_data.get('recycle_fee', 0)),
            'listing_fee': to_decimal(item_data.get('listing_fee', 0)),
            'listing_fee_tax': to_decimal(item_data.get('listing_fee_tax', 0)),
            'successful_bid': to_decimal(item_data.get('successful_bid', 0)),
            'successful_bid_tax': to_decimal(item_data.get('successful_bid_tax', 0)),
            'commission_fee': to_decimal(item_data.get('commission_fee', 0)),
            'commission_fee_tax': to_decimal(item_data.get('commission_fee_tax', 0)),
            'transport_fee': to_decimal(item_data.get('transport_fee', 0)),
            'transport_fee_tax': to_decimal(item_data.get('transport_fee_tax', 0)),
            'registration_fee': to_decimal(item_data.get('registration_fee', 0)),
            'registration_fee_tax': to_decimal(item_data.get('registration_fee_tax', 0)),
            'canceling_fee': to_decimal(item_data.get('canceling_fee', 0)),
        }

    def _resolve_category_for_item(self, user, raw_category):
        category_value = str(raw_category or '').strip()
        if not category_value:
            raise ValueError('Car name/category is required')

        if category_value.isdigit():
            category = CarCategory.objects.filter(id=int(category_value), user=user).first()
            if category:
                return category

        category = CarCategory.objects.filter(
            user=user
        ).filter(
            Q(name__iexact=category_value) | Q(company__iexact=category_value)
        ).first()
        if category:
            return category

        category = CarCategory.objects.filter(name=category_value, company=category_value).first()
        if category:
            return category

        try:
            return CarCategory.objects.create(
                user=user,
                name=category_value,
                company=category_value,
                description=''
            )
        except IntegrityError:
            category = CarCategory.objects.filter(name=category_value, company=category_value).first()
            if category:
                return category
            raise

    @action(detail=False, methods=['post'])
    def create_with_items(self, request):
        serializer = CreateOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        items_data = data.pop('items')
        customer_id = data.pop('customer_id', None)
        saler_id = data.pop('saler_id', None)
        company_account_id = data.pop('company_account_id', None)
        auction_id = data.pop('auction_id', None)
        transaction_id = data.pop('transaction', None)

        customer = None
        if customer_id:
            try:
                customer = Customer.objects.get(id=customer_id, user=request.user)
            except Customer.DoesNotExist:
                return Response({'error': 'Customer not found'}, status=status.HTTP_400_BAD_REQUEST)

        saler = None
        if saler_id:
            try:
                saler = Saler.objects.get(id=saler_id, user=request.user)
            except Saler.DoesNotExist:
                return Response({'error': 'Saler not found'}, status=status.HTTP_400_BAD_REQUEST)

        company_account = None
        if company_account_id:
            try:
                company_account = CompanyAccount.objects.get(id=company_account_id, user=request.user)
            except CompanyAccount.DoesNotExist:
                return Response({'error': 'Company account not found'}, status=status.HTTP_400_BAD_REQUEST)

        auction = None
        if auction_id:
            try:
                auction = Auction.objects.get(id=auction_id, user=request.user)
            except Auction.DoesNotExist:
                return Response({'error': 'Auction not found'}, status=status.HTTP_400_BAD_REQUEST)

        transaction_obj = None
        if transaction_id:
            try:
                transaction_obj = Transaction.objects.get(id=transaction_id, user=request.user)
            except Transaction.DoesNotExist:
                return Response({'error': 'Transaction not found'}, status=status.HTTP_400_BAD_REQUEST)

        if data['transaction_type'] == 'sale' and not customer:
            return Response({'error': 'Customer is required for sale orders'}, status=status.HTTP_400_BAD_REQUEST)
        if data['transaction_type'] == 'purchase' and not saler:
            return Response({'error': 'Saler is required for purchase orders'}, status=status.HTTP_400_BAD_REQUEST)
        if data['transaction_type'] == 'auction' and not customer:
            return Response({'error': 'Customer is required for auction orders'}, status=status.HTTP_400_BAD_REQUEST)

        other_details = {
            'customer_name': data.pop('customer_name', ''),
            'saler_name': data.pop('saler_name', ''),
            'seller_name': data.pop('seller_name', ''),
            'phone': data.pop('phone', ''),
            'address': data.pop('address', ''),
            'payment_method': data.pop('payment_method', ''),
            'account_number': data.pop('account_number', ''),
            'my_payment_method': data.pop('my_payment_method', ''),
            'my_account_number': data.pop('my_account_number', ''),
            'auction_house': data.pop('auction_house', '')
        }

        resolved_categories = []
        for item_data in items_data:
            try:
                resolved_categories.append(self._resolve_category_for_item(request.user, item_data.get('category')))
            except ValueError as exc:
                return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        for item_data, category in zip(items_data, resolved_categories):
                car_model = item_data.get('model') or category.name
                car_e = Car.objects.filter(chassis_number=item_data['chassis_number']).first()
                if car_e:
                    return Response({'error': f"Car with chassis number {item_data['chassis_number']} already exists"}, status=status.HTTP_400_BAD_REQUEST)

        today = datetime.now().date()
        date_str = today.strftime('%Y%m%d')
        last_order = Order.objects.filter(order_number__startswith=f'ORD-{date_str}').order_by('-order_number').first()
        new_num = int(last_order.order_number.split('-')[-1]) + 1 if last_order else 1
        order_number = f'ORD-{date_str}-{new_num:03d}'

        total_amount = sum(self._calculate_item_total(item) for item in items_data)

        with transaction.atomic():
            order = Order.objects.create(
                user=request.user,
                order_number=order_number,
                total_amount=total_amount,
                customer_name=other_details.get('customer_name') if data['transaction_type'] in ['sale', 'auction', 'nagare'] else other_details.get('saler_name', ''),
                other_details=other_details,
                transaction_type=data['transaction_type'],
                transaction_catagory=data['transaction_catagory'],
                transaction_date=data['transaction_date'],
                payment_status=data['payment_status'],
                notes=data.get('notes', ''),
                customer=customer if data['transaction_type'] in ['sale', 'auction', 'nagare'] else None,
                saler=saler if data['transaction_type'] == 'purchase' else None,
                company_account=company_account,
                auction=auction,
                transaction=transaction_obj
            )

            for item_data, category in zip(items_data, resolved_categories):
                car_model = item_data.get('model') or category.name
                car= Car.objects.create(
                    user=request.user,
                    category=category,
                    model=car_model,
                    chassis_number=item_data['chassis_number'],
                    year=item_data['year']
                )

                OrderItem.objects.create(
                    order=order,
                    car=car,
                    car_category=category,
                    **self._build_order_item_payload(item_data)
                )

        return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def update_with_items(self, request, pk=None):
        order = self.get_object()
        serializer = CreateOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        items_data = data.pop('items')
        customer_id = data.pop('customer_id', None)
        saler_id = data.pop('saler_id', None)
        company_account_id = data.pop('company_account_id', None)
        auction_id = data.pop('auction_id', None)
        transaction_id = data.pop('transaction', None)

        customer = None
        if customer_id:
            try:
                customer = Customer.objects.get(id=customer_id, user=request.user)
            except Customer.DoesNotExist:
                return Response({'error': 'Customer not found'}, status=status.HTTP_400_BAD_REQUEST)

        saler = None
        if saler_id:
            try:
                saler = Saler.objects.get(id=saler_id, user=request.user)
            except Saler.DoesNotExist:
                return Response({'error': 'Saler not found'}, status=status.HTTP_400_BAD_REQUEST)

        company_account = None
        if company_account_id:
            try:
                company_account = CompanyAccount.objects.get(id=company_account_id, user=request.user)
            except CompanyAccount.DoesNotExist:
                return Response({'error': 'Company account not found'}, status=status.HTTP_400_BAD_REQUEST)

        auction = None
        if auction_id:
            try:
                auction = Auction.objects.get(id=auction_id, user=request.user)
            except Auction.DoesNotExist:
                return Response({'error': 'Auction not found'}, status=status.HTTP_400_BAD_REQUEST)

        transaction_obj = None
        if transaction_id:
            try:
                transaction_obj = Transaction.objects.get(id=transaction_id, user=request.user)
            except Transaction.DoesNotExist:
                return Response({'error': 'Transaction not found'}, status=status.HTTP_400_BAD_REQUEST)

        other_details = {
            'customer_name': data.pop('customer_name', ''),
            'saler_name': data.pop('saler_name', ''),
            'seller_name': data.pop('seller_name', ''),
            'phone': data.pop('phone', ''),
            'address': data.pop('address', ''),
            'payment_method': data.pop('payment_method', ''),
            'account_number': data.pop('account_number', ''),
            'my_payment_method': data.pop('my_payment_method', ''),
            'my_account_number': data.pop('my_account_number', ''),
            'auction_house': data.pop('auction_house', '')
        }

        resolved_categories = []
        for item_data in items_data:
            try:
                resolved_categories.append(self._resolve_category_for_item(request.user, item_data.get('category')))
            except ValueError as exc:
                return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        total_amount = sum(self._calculate_item_total(item) for item in items_data)

        with transaction.atomic():
            order.total_amount = total_amount
            order.customer_name = other_details.get('customer_name') if data['transaction_type'] in ['sale', 'auction', 'nagare'] else other_details.get('saler_name', '')
            order.other_details = other_details
            order.transaction_type = data['transaction_type']
            order.transaction_catagory = data['transaction_catagory']
            order.transaction_date = data['transaction_date']
            order.payment_status = data['payment_status']
            order.notes = data.get('notes', '')
            order.customer = customer if data['transaction_type'] in ['sale', 'auction', 'nagare'] else None
            order.saler = saler if data['transaction_type'] == 'purchase' else None
            order.company_account = company_account
            order.auction = auction
            order.transaction = transaction_obj
            order.save()

            order.items.all().delete()

            for item_data, category in zip(items_data, resolved_categories):
                car_model = item_data.get('model') or category.name
                car, _ = Car.objects.get_or_create(
                    user=request.user,
                    category=category,
                    model=car_model,
                    chassis_number=item_data['chassis_number'],
                    year=item_data['year']
                )

                OrderItem.objects.create(
                    order=order,
                    car=car,
                    car_category=category,
                    **self._build_order_item_payload(item_data)
                )

        return Response(OrderSerializer(order).data)
        serializer = CreateOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        items_data = data.pop('items')
        customer_id = data.pop('customer_id', None)
        saler_id = data.pop('saler_id', None)
        company_account_id = data.pop('company_account_id', None)
        auction_id = data.pop('auction_id', None)
        transaction_id = data.pop('transaction', None)

        # Validate foreign keys
        customer = None
        if customer_id:
            try:
                customer = Customer.objects.get(id=customer_id, user=request.user)
            except Customer.DoesNotExist:
                return Response({'error': 'Customer not found'}, status=status.HTTP_400_BAD_REQUEST)

        saler = None
        if saler_id:
            try:
                saler = Saler.objects.get(id=saler_id, user=request.user)
            except Saler.DoesNotExist:
                return Response({'error': 'Saler not found'}, status=status.HTTP_400_BAD_REQUEST)

        company_account = None
        if company_account_id:
            try:
                company_account = CompanyAccount.objects.get(id=company_account_id, user=request.user)
            except CompanyAccount.DoesNotExist:
                return Response({'error': 'Company account not found'}, status=status.HTTP_400_BAD_REQUEST)

        auction = None
        if auction_id:
            try:
                auction = Auction.objects.get(id=auction_id, user=request.user)
            except Auction.DoesNotExist:
                return Response({'error': 'Auction not found'}, status=status.HTTP_400_BAD_REQUEST)

        transaction_obj = None
        if transaction_id:
            try:
                transaction_obj = Transaction.objects.get(id=transaction_id, user=request.user)
            except Transaction.DoesNotExist:
                return Response({'error': 'Transaction not found'}, status=status.HTTP_400_BAD_REQUEST)

        if data['transaction_type'] == 'sale' and not customer:
            return Response({'error': 'Customer is required for sale orders'}, status=status.HTTP_400_BAD_REQUEST)
        if data['transaction_type'] == 'purchase' and not saler:
            return Response({'error': 'Saler is required for purchase orders'}, status=status.HTTP_400_BAD_REQUEST)
        if data['transaction_type'] == 'auction' and not customer:
            return Response({'error': 'Customer is required for auction orders'}, status=status.HTTP_400_BAD_REQUEST)

        # Move all extra fields into other_details
        other_details = {
            'customer_name': data.pop('customer_name', ''),
            'saler_name': data.pop('saler_name', ''),
            'seller_name': data.pop('seller_name', ''),
            'phone': data.pop('phone', ''),
            'address': data.pop('address', ''),
            'payment_method': data.pop('payment_method', ''),
            'account_number': data.pop('account_number', ''),
            'my_payment_method': data.pop('my_payment_method', ''),
            'my_account_number': data.pop('my_account_number', ''),
            'auction_house': data.pop('auction_house', '')
        }

        # Check categories exist
        category_ids = [item['category'] for item in items_data]
        existing_categories = CarCategory.objects.filter(id__in=category_ids, user=request.user).values_list('id', flat=True)
        missing_categories = set(category_ids) - set(existing_categories)
        if missing_categories:
            return Response(
                {'error': f'Categories with IDs {list(missing_categories)} do not exist or do not belong to you'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create order number
        today = datetime.now().date()
        date_str = today.strftime('%Y%m%d')
        last_order = Order.objects.filter(order_number__startswith=f'ORD-{date_str}').order_by('-order_number').first()
        new_num = int(last_order.order_number.split('-')[-1]) + 1 if last_order else 1
        order_number = f'ORD-{date_str}-{new_num:03d}'

        # Calculate total_amount
        total_amount = sum(self._calculate_item_total(item) for item in items_data)

        # Create order
        with transaction.atomic():
            order = Order.objects.create(
                user=request.user,
                order_number=order_number,
                total_amount=total_amount,
                customer_name=other_details.get('customer_name') if data['transaction_type'] in ['sale', 'auction'] else other_details.get('saler_name', ''),
                other_details=other_details,
                transaction_type=data['transaction_type'],
                transaction_catagory=data['transaction_catagory'],
                transaction_date=data['transaction_date'],
                payment_status=data['payment_status'],
                notes=data.get('notes', ''),
                customer=customer if data['transaction_type'] in ['sale', 'auction'] else None,
                saler=saler if data['transaction_type'] == 'purchase' else None,
                company_account=company_account,
                auction=auction,
                transaction=transaction_obj
            )

            for item_data in items_data:
                category = CarCategory.objects.get(id=item_data['category'])
                car, _ = Car.objects.get_or_create(
                    user=request.user,
                    category_id=item_data['category'],
                    model=item_data['model'],
                    chassis_number=item_data['chassis_number'],
                    year=item_data['year']
                )

                OrderItem.objects.create(
                    order=order,
                    car=car,
                    car_category=category,
                    **self._build_order_item_payload(item_data)
                )

        return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)

    def _add_watermark(self, canvas, doc, text="INVOICE"):
        canvas.saveState()
        canvas.setFont("HeiseiMin-W3", 80)
        canvas.setFillColor(colors.lightgrey)
        canvas.setFillAlpha(0.15)

        width, height = doc.pagesize
        canvas.translate(width / 2, height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, text)

        canvas.restoreState()

    @action(detail=True, methods=['get'])
    def generate_invoice(self, request, pk=None):
        order = self.get_object()
        is_auction = order.transaction_type == 'auction'
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        user = User.objects.filter(email='user@example.com').first()
        if not user:
            return Response({'error': 'Admin user not found'}, status=404)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="Invoice_{order.order_number}.pdf"'
        )

        pagesize = landscape(A4)
        # pagesize = landscape(A4) if is_auction else A4

        doc = SimpleDocTemplate(
            response,
            pagesize=pagesize,
            topMargin=15,
            bottomMargin=15,
            leftMargin=10,
            rightMargin=10
        )

        elements = []
        styles = getSampleStyleSheet()

        # =====================================================
        # HEADER WITH COMPANY INFO AND INVOICE DETAILS
        # =====================================================
        header_elements = self._build_header_section(order, user, doc, styles)
        elements.extend(header_elements)
        elements.append(Spacer(1, 10))
        
        # =====================================================
        # HORIZONTAL LINE
        # =====================================================
        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1, lineCap='round', color=colors.black))
        elements.append(Spacer(1, 15))

        # =====================================================
        # BANK AND CUSTOMER INFO SECTION
        # =====================================================
        elements.append(self._build_bank_customer_section(order, doc, styles))
        elements.append(Spacer(1, 20))

        # =====================================================
        # ITEMS TABLE
        # =====================================================
        if is_auction:
            elements.append(self._build_auction_table(order, doc, styles))
        else:
            elements.append(self._build_auction_table(order, doc, styles))
            # elements.append(self._build_standard_table(order, doc, styles))

        doc.build(
            elements,
            onFirstPage=self._add_page_decorations,
            onLaterPages=self._add_page_decorations,
        )

        return response

    def _add_page_decorations(self, canvas, doc):
        from django.conf import settings
        import os
        from reportlab.lib.utils import ImageReader
        self._add_watermark(canvas, doc, "Ilyas Sons 合同会社")

        logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")
        print(logo_path)
        if os.path.exists(logo_path):
            logo = ImageReader(logo_path)

            logo_width = 120
            logo_height = 40

            x = doc.leftMargin
            y = doc.pagesize[1] - logo_height - 10

            canvas.drawImage(
                logo,
                x,
                y,
                width=logo_width,
                height=logo_height,
                mask='auto'
            )

        canvas.saveState()
        canvas.setFont("HeiseiMin-W3", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(
            doc.pagesize[0] / 2,
            12,
            "コンピュータで生成された請求書です。"
        )
        canvas.restoreState()
    
    # ======================================================
    # HEADER SECTION WITH COMPANY AND INVOICE INFO
    # ======================================================
    def _build_header_section(self, order, user, doc, styles):
        od = order.other_details or {}
        
        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Normal'],
            fontSize=24,
            alignment=TA_CENTER,
            fontName='HeiseiMin-W3'
        )
        title = Paragraph("請求書", title_style)
        
        # Company info in top right, invoice details below
        header_data = [
            ["", "", user.company_name],
            ["", "", user.company_address],
            ["", "", f"TEL/FAX: {user.company_phone}"],
            ["", "", f"{user.business_registration}"],
            ["", "", f"NO. {order.order_number}"],
            ["", "", f"Date : {order.transaction_date.strftime('%Y-%m-%d')}"],
        ]
        
        # Add transaction ID if available
        transaction_id = od.get('transaction_id', '')
        if transaction_id:
            header_data.append(["", "", f"取引ID: {transaction_id}"])
        
        header_table = Table(header_data, colWidths=[doc.width * 0.4, doc.width * 0.2, doc.width * 0.4])
        header_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (2, 0), (2, 3), 'HeiseiMin-W3'),
            ('FONTSIZE', (2, 4), (2, -1), 9),
        ]))
        
        return [title, Spacer(1, 15), header_table]

    # ======================================================
    # BANK AND CUSTOMER INFO SECTION
    # ======================================================
    def _build_bank_customer_section(self, order, doc, styles):
        if order.transaction_type == 'purchase':
            left_data = self._get_company_bank_info(order)
            middle_data = self._get_additional_info(order, styles)
            right_data = self._get_saler_info(order)
        else:
            left_data = self._get_company_bank_info(order)
            middle_data = self._get_additional_info(order, styles)
            right_data = self._get_customer_info(order)
        
        section_data = []
        max_rows = max(len(left_data), len(middle_data), len(right_data))
        
        for i in range(max_rows):
            left_cell = left_data[i] if i < len(left_data) else ""
            middle_cell = middle_data[i] if i < len(middle_data) else ""
            right_cell = right_data[i] if i < len(right_data) else ""
            section_data.append([left_cell, middle_cell, right_cell])
        
        section_table = Table(section_data, colWidths=[doc.width * 0.33, doc.width * 0.33, doc.width * 0.34])
        section_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'HeiseiMin-W3'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        return section_table
    
    def _get_additional_info(self, order, styles):
        data = []
        
        if order.auction:
            data.append(f"オークションハウス: {order.auction.name}")
        
        if order.transaction:
            data.append(f"取引ID: {order.transaction.transaction_id or order.transaction.description}")
        
        # Payment status
        status_map = {'pending': '保留中', 'completed': '完了', 'failed': '失敗'}
        payment_status = status_map.get(order.payment_status, order.payment_status)
        data.append(f"支払状況: {payment_status}")
        
        grand_total = sum(item.subtotal for item in order.items.all())
        total_style = ParagraphStyle(
            'TotalAmount',
            parent=styles['Normal'],
            fontSize=14,
            fontName='HeiseiMin-W3',
            textColor=colors.black,
            wordWrap='CJK'
        )
        total_para = Paragraph(f"合計金額 ¥ {grand_total:,.0f}", total_style)
        data.append(total_para)
        
        return data
    
    def _get_company_bank_info(self, order):
        od = order.other_details or {}
        payment_method = od.get('payment_method', 'Cash')
        
        data = []
        method = '銀行' if payment_method == 'Bank' else '現金'
        data.append(f"支払方法: {method}")
        if payment_method == 'Bank' and order.company_account:
            data.append(f"銀行: {order.company_account.bank_name}")
            data.append(f"普通: {order.company_account.account_number}")
            if order.company_account.branch_code:
                data.append(f"支店: {order.company_account.branch_code}")
            data.append(f"株式会社 {order.company_account.bank_name}")
        return data
    
    def _get_customer_info(self, order):
        od = order.other_details or {}
        payment_method = od.get('payment_method', 'Cash')
        data = []
        
        if order.customer:
            data.append(f"顧客の住所: {order.customer.address}" or "")
            data.append(f"顧客名: {order.customer.name}" or "")
            if payment_method == 'Bank' and hasattr(order.customer, 'bank_name') and order.customer.bank_name:
                data.append(f"銀行: {order.customer.bank_name}")
                data.append(f"口座: {order.customer.account_number}")
        else:
            data.append(od.get('address', ''))
            data.append(od.get('customer_name', ''))
        
        return data
    
    def _get_saler_info(self, order):
        od = order.other_details or {}
        payment_method = od.get('payment_method', 'Cash')
        data = []
        
        if order.saler:
            data.append(f"販売者の住所: {order.saler.address}" or "")
            data.append(f"販売者名: {order.saler.name}" or "")
            if payment_method == 'Bank' and hasattr(order.saler, 'bank_name') and order.saler.bank_name:
                data.append(f"銀行: {order.saler.bank_name}")
                data.append(f"口座: {order.saler.account_number}")
        else:
            data.append(od.get('address', ''))
            data.append(od.get('saler_name', ''))
        
        return data

    # ======================================================
    # AUCTION TABLE
    # ======================================================

    def _build_auction_table(self, order, doc, styles):
        header = [
    'No.',              # NO.
    '会場',              # Venue
    'カテゴリー',        # Category
    'モデル',            # Model
    '年',            # Year
    '車台番号',          # Chassis
    '車両価格',          # Vehicle Price
    'リサイクル料金',    # Recycle Fee
    '出品料',            # Listing Fee
    '落札料',            # Successful Bid
    '手数料',            # Commission Fee
    '輸送料',            # Transport Fee
    '登録料',            # Registration Fee
    'キャンセル料',      # Canceling Fee
    '合計'               # Total
]

        data = [header]
        totals = {
            'vehicle_price': 0, 'vehicle_price_tax': 0,
            'recycle_fee': 0, 'listing_fee': 0, 'listing_fee_tax': 0,
            'successful_bid': 0, 'successful_bid_tax': 0,
            'commission_fee': 0, 'commission_fee_tax': 0,
            'transport_fee': 0, 'transport_fee_tax': 0,
            'registration_fee': 0, 'registration_fee_tax': 0,
            'canceling_fee': 0, 'subtotal': 0,
        }
        small_style = ParagraphStyle(
            'SmallTable',
            parent=styles['Normal'],
            fontSize=6,
            leading=7,
            fontName='HeiseiMin-W3'
        )
        for idx, item in enumerate(order.items.all(), 1):
            row = [
                str(idx),
                Paragraph(item.venue or '', small_style),
                Paragraph(str(item.car_category.name) if item.car_category else '', small_style),
                Paragraph(item.car.model or '', small_style),
                Paragraph(item.car.year or '', small_style),
                Paragraph(item.car.chassis_number or '', small_style),
                Paragraph(f'{item.vehicle_price:,.0f}<br/>{item.vehicle_price_tax:,.0f}', small_style),
                Paragraph(f'{item.recycle_fee:,.0f}', small_style),
                Paragraph(f'{item.listing_fee:,.0f}<br/>{item.listing_fee_tax:,.0f}', small_style),
                Paragraph(f'{item.successful_bid:,.0f}<br/>{item.successful_bid_tax:,.0f}', small_style),
                Paragraph(f'{item.commission_fee:,.0f}<br/>{item.commission_fee_tax:,.0f}', small_style),
                Paragraph(f'{item.transport_fee:,.0f}<br/>{item.transport_fee_tax:,.0f}', small_style),
                Paragraph(f'{item.registration_fee:,.0f}<br/>{item.registration_fee_tax:,.0f}', small_style),
                Paragraph(f'{item.canceling_fee:,.0f}', small_style),
                Paragraph(f'{item.subtotal:,.0f}', small_style),
            ]
            data.append(row)

            totals['vehicle_price'] += item.vehicle_price
            totals['vehicle_price_tax'] += item.vehicle_price_tax
            totals['recycle_fee'] += item.recycle_fee
            totals['listing_fee'] += item.listing_fee
            totals['listing_fee_tax'] += item.listing_fee_tax
            totals['successful_bid'] += item.successful_bid
            totals['successful_bid_tax'] += item.successful_bid_tax
            totals['commission_fee'] += item.commission_fee
            totals['commission_fee_tax'] += item.commission_fee_tax
            totals['transport_fee'] += item.transport_fee
            totals['transport_fee_tax'] += item.transport_fee_tax
            totals['registration_fee'] += item.registration_fee
            totals['registration_fee_tax'] += item.registration_fee_tax
            totals['canceling_fee'] += item.canceling_fee
            totals['subtotal'] += item.subtotal

        data.append([
            '', '', '', '', '合計',
            Paragraph(f'{totals["vehicle_price"]:,.0f}<br/>{totals["vehicle_price_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["recycle_fee"]:,.0f}', small_style),
            Paragraph(f'{totals["listing_fee"]:,.0f}<br/>{totals["listing_fee_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["successful_bid"]:,.0f}<br/>{totals["successful_bid_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["commission_fee"]:,.0f}<br/>{totals["commission_fee_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["transport_fee"]:,.0f}<br/>{totals["transport_fee_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["registration_fee"]:,.0f}<br/>{totals["registration_fee_tax"]:,.0f}', small_style),
            Paragraph(f'{totals["canceling_fee"]:,.0f}', small_style),
            Paragraph(f'{totals["subtotal"]:,.0f}', small_style),
        ])
        base_col_widths = [18, 36, 40, 42, 42, 55, 45, 35, 45, 45, 45, 45, 45, 35, 45]
        base_total = sum(base_col_widths)
        target_width = doc.width * 0.99
        scale = target_width / base_total if base_total else 1
        col_widths = [w * scale for w in base_col_widths]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'HeiseiMin-W3'),
            ('FONTSIZE', (0, 0), (-1, 0), 5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (4, -2), 'CENTER'),
            ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -1), 4),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'HeiseiMin-W3'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))

        return table

    def _build_standard_table(self, order, doc, styles):
        header = ['NO.', '車種', 'モデル', 'シャーシ', '年式', '価格', '消費税', '合計']
        data = [header]
        totals = {'vehicle_price': 0, 'consumption_tax': 0, 'subtotal': 0}

        for idx, item in enumerate(order.items.all(), 1):
            data.append([
                str(idx), str(item.car.category) if item.car.category else '',
                item.car.model, item.car.chassis_number, str(item.car.year),
                f'{item.vehicle_price:,.0f}', f'{item.consumption_tax:,.0f}',
                f'{item.subtotal:,.0f}'
            ])
            
            totals['vehicle_price'] += item.vehicle_price
            totals['consumption_tax'] += item.consumption_tax
            totals['subtotal'] += item.subtotal

        data.append([
            '', '', '', '', '合計',
            f'{totals["vehicle_price"]:,.0f}',
            f'{totals["consumption_tax"]:,.0f}',
            f'{totals["subtotal"]:,.0f}'
        ])

        table = Table(data, colWidths=[30, 80, 100, 120, 50, 80, 80, 90], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'HeiseiMin-W3'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (4, -1), 'CENTER'),
            ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'HeiseiMin-W3'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        return table

    def _build_grand_total(self, order, styles):

        grand_total = sum(item.subtotal for item in order.items.all())

        style = ParagraphStyle(
            '総計',
            parent=styles['Normal'],
            fontSize=14,
            alignment=TA_RIGHT,
            fontName='HeiseiMin-W3'
        )

        return Paragraph(f" 総計 : ¥ {grand_total:,.0f}", style)


    @action(detail=False, methods=['get'])
    def financial_report(self, request):
        period = request.query_params.get('period', 'month')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        today = datetime.now().date()
        
        if start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        elif period == 'today':
            start = end = today
        elif period == 'month':
            start = today.replace(day=1)
            end = today
        elif period == 'year':
            start = today.replace(month=1, day=1)
            end = today
        else:
            start = today.replace(day=1)
            end = today
        
        orders = Order.objects.filter(user=request.user, transaction_date__range=[start, end])
        expenses = Expense.objects.filter(user=request.user, date__range=[start, end])
        
        sales = orders.filter(transaction_type='sale').aggregate(total=Sum('total_amount'))['total'] or 0
        purchases = orders.filter(transaction_type='purchase').aggregate(total=Sum('total_amount'))['total'] or 0
        auctions = orders.filter(transaction_type='auction').aggregate(total=Sum('total_amount'))['total'] or 0
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
        
        revenue = sales + auctions
        cost = purchases + total_expenses
        profit = revenue - cost
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Financial Report'
        
        ws['A1'] = 'Financial Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Period: {start} to {end}'
        
        ws['A4'] = 'Summary'
        ws['A4'].font = Font(bold=True)
        ws['A5'] = 'Currency'
        ws['B5'] = '¥'
        ws['A6'] = 'Total Revenue (Sales + Auctions)'
        ws['B6'] = f'¥ {float(revenue)}'
        ws['A7'] = 'Total Cost (Purchases + Expenses)'
        ws['B7'] = f'¥ {float(cost)}'
        ws['A8'] = 'Net Profit'
        ws['B8'] = f'¥ {float(profit)}'
        ws['B9'].font = Font(bold=True)
        
        ws['A10'] = 'Breakdown'
        ws['A10'].font = Font(bold=True)
        ws['A11'] = 'Sales'
        ws['B11'] = f'¥ {float(sales)}'
        ws['A12'] = 'Auctions'
        ws['B12'] = f'¥ {float(auctions)}'
        ws['A13'] = 'Purchases'
        ws['B13'] = f'¥ {float(purchases)}'
        ws['A14'] = 'Expenses'
        ws['B14'] = f'¥ {float(total_expenses)}'
        
        ws['A16'] = 'Transactions Detail'
        ws['A16'].font = Font(bold=True)
        ws.append(['Type', 'Date', 'Payment Status', 'Amount'])
        
        for order in orders:
            ws.append([order.transaction_type.capitalize(), order.transaction_date.strftime('%Y-%m-%d'), 
                      order.payment_status, f'¥ {float(order.total_amount)}'])
        
        for exp in expenses:
            ws.append(['Expense', exp.date.strftime('%Y-%m-%d'), 'completed', f'¥ {float(exp.amount)}'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Report {start}_{end}.xlsx"'
        wb.save(response)
        return response



    @action(detail=False, methods=['get'])
    def reports(self, request):
        from rest_framework.pagination import PageNumberPagination
        
        report_type = request.query_params.get('type', 'orders')
        period = request.query_params.get('period', 'month')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        payment_status = request.query_params.get('payment_status')
        search = request.query_params.get('search')
        page_size = int(request.query_params.get('pageSize', 10))
        
        today = datetime.now().date()
        
        if start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        elif period == 'today':
            start = end = today
        elif period == 'month':
            start = today.replace(day=1)
            end = today
        elif period == 'year':
            start = today.replace(month=1, day=1)
            end = today
        else:
            start = today.replace(day=1)
            end = today
        
        data = []
        
        if report_type in ['all', 'expenses']:
            if period == 'all':
                expenses = Expense.objects.filter(user=request.user)
            else:
                expenses = Expense.objects.filter(user=request.user, date__range=[start, end])
            if search:
                expenses = expenses.filter(Q(description__icontains=search))
            for exp in expenses:
                data.append({
                    'transaction_type': 'expense',
                    'transaction_date': exp.date,
                    'payment_status': 'completed',
                    'total_amount': exp.amount
                })
        
        if report_type in ['all', 'orders', 'sales', 'purchases', 'auctions']:
            if period == 'all':
                queryset = Order.objects.filter(user=request.user)
            else:
                queryset = Order.objects.filter(user=request.user, transaction_date__range=[start, end])
            
            if report_type == 'sales':
                queryset = queryset.filter(transaction_type='sale')
            elif report_type == 'purchases':
                queryset = queryset.filter(transaction_type='purchase')
            elif report_type == 'auctions':
                queryset = queryset.filter(transaction_type='auction')
            
            if payment_status:
                queryset = queryset.filter(payment_status=payment_status)
            if search:
                queryset = queryset.filter(Q(order_number__icontains=search) | Q(customer_name__icontains=search))
            
            for order in queryset.values('transaction_type', 'transaction_date', 'payment_status', 'total_amount'):
                data.append(order)
        
        data.sort(key=lambda x: x['transaction_date'], reverse=True)
        
        paginator = PageNumberPagination()
        paginator.page_size = page_size
        paginated = paginator.paginate_queryset(data, request)
        
        return paginator.get_paginated_response(paginated)


class OrderItemViewSet(viewsets.ModelViewSet):
    serializer_class = OrderItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return OrderItem.objects.filter(order__user=self.request.user)


class CustomerViewSet(viewsets.ModelViewSet):
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = Customer.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class SalerViewSet(viewsets.ModelViewSet):
    serializer_class = SalerSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = Saler.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class CompanyAccountViewSet(viewsets.ModelViewSet):
    serializer_class = CompanyAccountSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = CompanyAccount.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(bank_name__icontains=search) |
                Q(account_number__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class AuctionViewSet(viewsets.ModelViewSet):
    serializer_class = AuctionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = Auction.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = Transaction.objects.filter(user=self.request.user).order_by('-date', '-id')
        search = self.request.query_params.get('search')
        date = self.request.query_params.get('date')
        company_account = self.request.query_params.get('company_account')
        
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(notes__icontains=search) |
                Q(transaction_id__icontains=search) |
                Q(balance__icontains=search) |
                Q(deposit__icontains=search) |
                Q(withdraw__icontains=search)
            )
        if date:
            queryset = queryset.filter(date=date)
        if company_account:
            queryset = queryset.filter(company_account_id=company_account)
            
        return queryset

    def perform_create(self, serializer):
        # Calculate balance based on previous transactions for the same account
        company_account = serializer.validated_data.get('company_account')
        date = serializer.validated_data.get('date')
        withdraw = serializer.validated_data.get('withdraw', 0)
        deposit = serializer.validated_data.get('deposit', 0)
        
        # Get the latest transaction for this account before the current date
        latest_transaction = Transaction.objects.filter(
            user=self.request.user,
            company_account=company_account,
            date__lte=date
        ).order_by('-date', '-id').first()
        
        # Calculate new balance
        previous_balance = latest_transaction.balance if latest_transaction else 0
        new_balance = previous_balance + deposit - withdraw
        
        serializer.save(user=self.request.user, balance=new_balance)
        
        # Update balances for all subsequent transactions
        self._update_subsequent_balances(company_account, date, serializer.instance.id)
    
    def perform_update(self, serializer):
        # Recalculate balance for updated transaction
        company_account = serializer.validated_data.get('company_account')
        date = serializer.validated_data.get('date')
        withdraw = serializer.validated_data.get('withdraw', 0)
        deposit = serializer.validated_data.get('deposit', 0)
        
        # Get the latest transaction for this account before the current date
        latest_transaction = Transaction.objects.filter(
            user=self.request.user,
            company_account=company_account,
            date__lt=date
        ).order_by('-date', '-id').first()
        
        # Calculate new balance
        previous_balance = latest_transaction.balance if latest_transaction else 0
        new_balance = previous_balance + deposit - withdraw
        
        serializer.save(balance=new_balance)
        
        # Update balances for all subsequent transactions
        self._update_subsequent_balances(company_account, date, serializer.instance.id)
    
    def _update_subsequent_balances(self, company_account, from_date, exclude_id):
        """Update balances for all transactions after the given date"""
        subsequent_transactions = Transaction.objects.filter(
            user=self.request.user,
            company_account=company_account,
            date__gte=from_date
        ).exclude(id=exclude_id).order_by('date', 'id')
        
        # Get the balance from the transaction just before the first subsequent transaction
        if subsequent_transactions.exists():
            first_subsequent = subsequent_transactions.first()
            previous_transaction = Transaction.objects.filter(
                user=self.request.user,
                company_account=company_account,
                date__lt=first_subsequent.date
            ).order_by('-date', '-id').first()
            
            running_balance = previous_transaction.balance if previous_transaction else 0
            
            for transaction in subsequent_transactions:
                running_balance = running_balance + transaction.deposit - transaction.withdraw
                transaction.balance = running_balance
                transaction.save(update_fields=['balance'])

    # ############ japan post
    # @action(detail=False, methods=['post'])
    # def bulk_import(self, request):
    #     from datetime import datetime
    #     import csv
    #     from io import StringIO
    #     import requests
        
    #     csv_data = request.data.get('csv_data', '')
    #     sheet_url = request.data.get('sheet_url', '')
        
    #     if not csv_data and not sheet_url:
    #         return Response({'error': 'Either csv_data or sheet_url is required'}, status=400)
        
    #     try:
    #         # If Google Sheets URL is provided, fetch CSV data
    #         if sheet_url:
    #             # Handle published CSV URLs directly
    #             if 'pub?output=csv' in sheet_url:
    #                 csv_url = sheet_url
    #             else:
    #                 # Extract sheet ID from various Google Sheets URL formats
    #                 sheet_id = None
    #                 gid = '0'  # default sheet
                    
    #                 if '/d/' in sheet_url:
    #                     # Extract sheet ID from URL like: https://docs.google.com/spreadsheets/d/ID/edit
    #                     sheet_id = sheet_url.split('/d/')[1].split('/')[0]
    #                     if 'gid=' in sheet_url:
    #                         gid = sheet_url.split('gid=')[1].split('&')[0].split('#')[0]
                    
    #                 if not sheet_id:
    #                     return Response({'error': 'Invalid Google Sheets URL format'}, status=400)
                    
    #                 # Construct proper CSV export URL
    #                 csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'
                
    #             try:
    #                 response = requests.get(csv_url, timeout=30)
    #                 response.raise_for_status()
    #                 response.encoding = 'utf-8'  # Ensure proper UTF-8 encoding
    #                 csv_data = response.text
    #             except requests.exceptions.HTTPError as e:
    #                 if e.response.status_code == 400:
    #                     return Response({
    #                         'error': 'Sheet access denied. Please make your Google Sheet public: Share > Anyone with the link > Viewer'
    #                     }, status=400)
    #                 raise
            
    #         csv_file = StringIO(csv_data)
    #         reader = csv.reader(csv_file)
            
    #         # Skip header row
    #         next(reader, None)
            
    #         transactions = []
    #         for row in reader:
    #             if len(row) < 6:  # Skip incomplete rows
    #                 continue
                    
    #             # Parse date (年月日)
    #             date_str = row[0].strip()
    #             try:
    #                 # Try multiple date formats
    #                 for fmt in [
    #                         '%Y/%m/%d',
    #                         '%Y-%m-%d',
    #                         '%d/%m/%Y',
    #                         '%m/%d/%Y',
    #                         '%b. %d, %Y',
    #                         '%Y%m%d', 
    #                         '%m/%d/%y %H:%M',
    #                         '%m/%d/%Y %H:%M',
    #                     ]:

    #                     try:
    #                         date_obj = datetime.strptime(date_str, fmt).date()
    #                         break
    #                     except ValueError:
    #                         continue
    #                 else:
    #                     continue  # Skip if no format matches
    #             except:
    #                 pass
                
    #             # Parse withdraw (お引出し)
    #             withdraw = 0
    #             if row[3].strip():
    #                 try:
    #                     if not row[3].strip() == '-':
    #                         withdraw = float(row[3].strip())
    #                 except ValueError:
    #                     withdraw = 0
                
    #             # Parse deposit (お預入れ)
    #             deposit = 0
    #             if row[2].strip():
    #                 try:
    #                     if not row[2].strip() == '-':
    #                         deposit = float(row[2].strip())
    #                 except ValueError:
    #                     deposit = 0
                
    #             # Parse description (お取り扱い内容)
    #             description = row[5].strip()[:500]  # Limit to 500 chars
                
    #             # Parse balance (残高)
    #             balance = 0
    #             if row[6].strip():
    #                 try:
    #                     balance = float(row[6].strip())
    #                 except ValueError:
    #                     balance = 0
                
    #             # Skip メモ column (row[5]) as requested
    #             # Notes from ラベル column (row[6] if exists)
    #             notes = ''
    #             if len(row) > 3 and row[4].strip():
    #                 notes = row[4].strip()

    #             transactionId = ''
    #             if len(row) > 6 and row[1].strip():
    #                 transactionId = row[1].strip()
    #             company_account = None
    #             if len(row) > 5 and row[7].strip():
    #                 try:
    #                     account_id = int(row[7].strip())
    #                     company_account = CompanyAccount.objects.filter(id=account_id).first()
    #                 except ValueError:
    #                     company_account = None
    #             transactions.append(Transaction(
    #                 user=request.user,
    #                 date=date_obj,
    #                 withdraw=withdraw,
    #                 deposit=deposit,
    #                 balance=balance,
    #                 description=description,
    #                 notes=notes,
    #                 company_account=company_account,
    #                 transaction_id=transactionId
    #             ))
            
    #         # Bulk create transactions
    #         Transaction.objects.bulk_create(transactions, ignore_conflicts=True)
            
    #         return Response({
    #             'message': f'Successfully imported {len(transactions)} transactions',
    #             'count': len(transactions)
    #         })
            
    #     except requests.RequestException as e:
    #         return Response({'error': f'Failed to fetch data from URL: {str(e)}'}, status=400)
    #     except Exception as e:
    #         return Response({'error': str(e)}, status=400)
    

    ########### paypay
    # @action(detail=False, methods=['post'])
    # def bulk_import(self, request):
    #     from datetime import datetime
    #     import csv
    #     from io import StringIO
    #     import requests
        
    #     csv_data = request.data.get('csv_data', '')
    #     sheet_url = request.data.get('sheet_url', '')
        
    #     if not csv_data and not sheet_url:
    #         return Response({'error': 'Either csv_data or sheet_url is required'}, status=400)
        
    #     try:
    #         # If Google Sheets URL is provided, fetch CSV data
    #         if sheet_url:
    #             # Handle published CSV URLs directly
    #             if 'pub?output=csv' in sheet_url:
    #                 csv_url = sheet_url
    #             else:
    #                 # Extract sheet ID from various Google Sheets URL formats
    #                 sheet_id = None
    #                 gid = '0'  # default sheet
                    
    #                 if '/d/' in sheet_url:
    #                     # Extract sheet ID from URL like: https://docs.google.com/spreadsheets/d/ID/edit
    #                     sheet_id = sheet_url.split('/d/')[1].split('/')[0]
    #                     if 'gid=' in sheet_url:
    #                         gid = sheet_url.split('gid=')[1].split('&')[0].split('#')[0]
                    
    #                 if not sheet_id:
    #                     return Response({'error': 'Invalid Google Sheets URL format'}, status=400)
                    
    #                 # Construct proper CSV export URL
    #                 csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'
                
    #             try:
    #                 response = requests.get(csv_url, timeout=30)
    #                 response.raise_for_status()
    #                 response.encoding = 'utf-8'  # Ensure proper UTF-8 encoding
    #                 csv_data = response.text
    #             except requests.exceptions.HTTPError as e:
    #                 if e.response.status_code == 400:
    #                     return Response({
    #                         'error': 'Sheet access denied. Please make your Google Sheet public: Share > Anyone with the link > Viewer'
    #                     }, status=400)
    #                 raise
            
    #         csv_file = StringIO(csv_data)
    #         reader = csv.reader(csv_file)
            
    #         # Skip header row
    #         next(reader, None)
            
    #         transactions = []
    #         for row in reader:
    #             if len(row) < 6:  # Skip incomplete rows
    #                 continue
                    
    #             # Parse date (年月日)
    #             date_str = row[0].strip()
    #             try:
    #                 # Try multiple date formats
    #                 for fmt in [
    #                         '%Y/%m/%d',
    #                         '%Y-%m-%d',
    #                         '%d/%m/%Y',
    #                         '%m/%d/%Y',
    #                         '%b. %d, %Y',
    #                         '%Y%m%d', 
    #                         '%m/%d/%y %H:%M',
    #                         '%m/%d/%Y %H:%M',
    #                     ]:

    #                     try:
    #                         date_obj = datetime.strptime(date_str, fmt).date()
    #                         break
    #                     except ValueError:
    #                         continue
    #                 else:
    #                     continue  # Skip if no format matches
    #             except:
    #                 pass
                
    #             # Parse withdraw (お引出し)
    #             withdraw = 0
    #             if row[1].strip():
    #                 try:
    #                     if not row[1].strip() == '-':
    #                         withdraw = float(row[1].strip())
    #                 except ValueError:
    #                     withdraw = 0
                
    #             # Parse deposit (お預入れ)
    #             deposit = 0
    #             if row[2].strip():
    #                 try:
    #                     if not row[2].strip() == '-':
    #                         deposit = float(row[2].strip())
    #                 except ValueError:
    #                     deposit = 0
                
    #             # Parse description (お取り扱い内容)
    #             description = row[8].strip()[:500]  # Limit to 500 chars
                
    #             # Parse balance (残高)
    #             balance = 0
    #             # if row[4].strip():
    #             #     try:
    #             #         balance = float(row[4].strip())
    #             #     except ValueError:
    #             #         balance = 0
                
    #             # Skip メモ column (row[5]) as requested
    #             # Notes from ラベル column (row[6] if exists)
    #             notes = ''
    #             # if len(row) > 6 and row[6].strip():
    #             #     notes = row[6].strip()

    #             transactionId = ''
    #             if len(row) > 6 and row[12].strip():
    #                 transactionId = row[12].strip()
    #             company_account = None
    #             if len(row) > 5 and row[13].strip():
    #                 try:
    #                     account_id = int(row[13].strip())
    #                     company_account = CompanyAccount.objects.filter(id=account_id).first()
    #                 except ValueError:
    #                     company_account = None
    #             transactions.append(Transaction(
    #                 user=request.user,
    #                 date=date_obj,
    #                 withdraw=withdraw,
    #                 deposit=deposit,
    #                 balance=balance,
    #                 description=description,
    #                 notes=notes,
    #                 company_account=company_account,
    #                 transaction_id=transactionId
    #             ))
            
    #         # Bulk create transactions
    #         Transaction.objects.bulk_create(transactions, ignore_conflicts=True)
            
    #         return Response({
    #             'message': f'Successfully imported {len(transactions)} transactions',
    #             'count': len(transactions)
    #         })
            
    #     except requests.RequestException as e:
    #         return Response({'error': f'Failed to fetch data from URL: {str(e)}'}, status=400)
    #     except Exception as e:
    #         return Response({'error': str(e)}, status=400)


    ####################### smbc
    # @action(detail=False, methods=['post'])
    # def bulk_import(self, request):
    #     from datetime import datetime
    #     import csv
    #     from io import StringIO
    #     import requests
    #     import re

    #     csv_data = request.data.get('csv_data', '')
    #     sheet_url = request.data.get('sheet_url', '')

    #     if not csv_data and not sheet_url:
    #         return Response({'error': 'Either csv_data or sheet_url is required'}, status=400)

    #     try:
    #         # ==============================
    #         # Fetch CSV from Google Sheets
    #         # ==============================
    #         if sheet_url:
    #             if 'pub?output=csv' in sheet_url:
    #                 csv_url = sheet_url
    #             else:
    #                 sheet_id = None
    #                 gid = '0'

    #                 if '/d/' in sheet_url:
    #                     sheet_id = sheet_url.split('/d/')[1].split('/')[0]
    #                     if 'gid=' in sheet_url:
    #                         gid = sheet_url.split('gid=')[1].split('&')[0].split('#')[0]

    #                 if not sheet_id:
    #                     return Response({'error': 'Invalid Google Sheets URL format'}, status=400)

    #                 csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'

    #             response = requests.get(csv_url, timeout=30)
    #             response.raise_for_status()
    #             response.encoding = 'utf-8'
    #             csv_data = response.text

    #         csv_file = StringIO(csv_data)
    #         reader = csv.reader(csv_file)

    #         # skip header
    #         next(reader, None)

    #         transactions = []

    #         for row in reader:
    #             # Expecting at least 6 columns
    #             if len(row) < 6:
    #                 continue

    #             # ==============================
    #             # DATE
    #             # ==============================
    #             date_str = row[0].strip()
    #             date_obj = None

    #             for fmt in (
    #                 '%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y',
    #                 '%m/%d/%Y', '%Y%m%d'
    #             ):
    #                 try:
    #                     date_obj = datetime.strptime(date_str, fmt).date()
    #                     break
    #                 except ValueError:
    #                     continue

    #             if not date_obj:
    #                 continue

    #             # ==============================
    #             # WITHDRAW
    #             # ==============================
    #             withdraw = 0
    #             if row[1].strip() and row[1].strip() != '-':
    #                 try:
    #                     withdraw = float(row[1].strip())
    #                 except:
    #                     pass

    #             # ==============================
    #             # DEPOSIT
    #             # ==============================
    #             deposit = 0
    #             if row[2].strip() and row[2].strip() != '-':
    #                 try:
    #                     deposit = float(row[2].strip())
    #                 except:
    #                     pass

    #             # ==============================
    #             # DESCRIPTION (お取り扱い内容)
    #             # ==============================
    #             transaction_id = row[3].strip()

    #             # normalize full-width spaces
    #             transaction_id = transaction_id.replace('\u3000', ' ')

    #             # extract codes like V495093
    #             match = re.match(r'^([A-Z]\d+)', transaction_id)

    #             if match:
    #                 transaction_id = match.group(1)
    #             else:
    #                 transaction_id = transaction_id

    #             transaction_id = transaction_id[:500]

    #             # ==============================
    #             # BALANCE
    #             # ==============================
    #             balance = 0
    #             if row[4].strip():
    #                 try:
    #                     balance = float(row[4].strip())
    #                 except:
    #                     pass
    #             company_account = None
    #             if len(row) > 5 and row[6].strip():
    #                 try:
    #                     account_id = int(row[6].strip())
    #                     company_account = CompanyAccount.objects.filter(id=account_id).first()
    #                 except ValueError:
    #                     company_account = None

    #             # ==============================
    #             # NOTES (メモ)
    #             # ==============================
    #             notes = ''
    #             # notes = row[5].strip() if len(row) > 5 else ''

    #             transactions.append(Transaction(
    #                 user=request.user,
    #                 date=date_obj,
    #                 withdraw=withdraw,
    #                 deposit=deposit,
    #                 balance=balance,
    #                 description='',
    #                 transaction_id=transaction_id,
    #                 notes=notes,
    #                 company_account=company_account,
    #             ))

    #         Transaction.objects.bulk_create(transactions)

    #         return Response({
    #             'message': f'Successfully imported {len(transactions)} transactions',
    #             'count': len(transactions)
    #         })

    #     except requests.RequestException as e:
    #         return Response({'error': f'Failed to fetch data: {str(e)}'}, status=400)
    #     except Exception as e:
    #         return Response({'error': str(e)}, status=400)

    ####################### mesai
    # @action(detail=False, methods=['post'])
    # def bulk_import(self, request):
    #     from datetime import datetime
    #     import csv
    #     from io import StringIO
    #     import requests
    #     import re

    #     csv_data = request.data.get('csv_data', '')
    #     sheet_url = request.data.get('sheet_url', '')

    #     if not csv_data and not sheet_url:
    #         return Response({'error': 'Either csv_data or sheet_url is required'}, status=400)

    #     try:
    #         # ==============================
    #         # Fetch CSV from Google Sheets
    #         # ==============================
    #         if sheet_url:
    #             if 'pub?output=csv' in sheet_url:
    #                 csv_url = sheet_url
    #             else:
    #                 sheet_id = None
    #                 gid = '0'

    #                 if '/d/' in sheet_url:
    #                     sheet_id = sheet_url.split('/d/')[1].split('/')[0]
    #                     if 'gid=' in sheet_url:
    #                         gid = sheet_url.split('gid=')[1].split('&')[0].split('#')[0]

    #                 if not sheet_id:
    #                     return Response({'error': 'Invalid Google Sheets URL format'}, status=400)

    #                 csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'

    #             response = requests.get(csv_url, timeout=30)
    #             response.raise_for_status()
    #             response.encoding = 'utf-8'
    #             csv_data = response.text

    #         csv_file = StringIO(csv_data)
    #         reader = csv.reader(csv_file)

    #         # skip header
    #         next(reader, None)

    #         transactions = []

    #         for row in reader:
    #             # Expecting at least 6 columns
    #             if len(row) < 6:
    #                 continue

    #             # ==============================
    #             # DATE
    #             # ==============================
    #             date_str = row[0].strip()
    #             date_obj = None

    #             for fmt in (
    #                 '%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y',
    #                 '%m/%d/%Y', '%Y%m%d'
    #             ):
    #                 try:
    #                     date_obj = datetime.strptime(date_str, fmt).date()
    #                     break
    #                 except ValueError:
    #                     continue

    #             if not date_obj:
    #                 continue

    #             # ==============================
    #             # WITHDRAW
    #             # ==============================
    #             withdraw = 0
    #             if row[1].strip() and row[1].strip() != '-':
    #                 try:
    #                     withdraw = float(row[1].strip())
    #                 except:
    #                     pass

    #             # ==============================
    #             # DEPOSIT
    #             # ==============================
    #             deposit = 0
    #             if row[2].strip() and row[2].strip() != '-':
    #                 try:
    #                     deposit = float(row[2].strip())
    #                 except:
    #                     pass

    #             # ==============================
    #             # DESCRIPTION (お取り扱い内容)
    #             # ==============================
    #             transaction_id = row[3].strip()

    #             # normalize full-width spaces
    #             transaction_id = transaction_id.replace('\u3000', ' ')

    #             # extract codes like V495093
    #             match = re.match(r'^([A-Z]\d+)', transaction_id)

    #             if match:
    #                 transaction_id = match.group(1)
    #             else:
    #                 transaction_id = transaction_id

    #             transaction_id = transaction_id[:500]

    #             # ==============================
    #             # BALANCE
    #             # ==============================
    #             balance = 0
    #             if row[4].strip():
    #                 try:
    #                     balance = float(row[4].strip())
    #                 except:
    #                     pass
    #             company_account = None
    #             if len(row) > 5 and row[5].strip():
    #                 try:
    #                     account_id = int(row[5].strip())
    #                     company_account = CompanyAccount.objects.filter(id=account_id).first()
    #                 except ValueError:
    #                     company_account = None

    #             # ==============================
    #             # NOTES (メモ)
    #             # ==============================
    #             notes = ''
    #             # notes = row[5].strip() if len(row) > 5 else ''

    #             transactions.append(Transaction(
    #                 user=request.user,
    #                 date=date_obj,
    #                 withdraw=withdraw,
    #                 deposit=deposit,
    #                 balance=balance,
    #                 description='',
    #                 transaction_id=transaction_id,
    #                 notes=notes,
    #                 company_account=company_account,
    #             ))

    #         Transaction.objects.bulk_create(transactions)

    #         return Response({
    #             'message': f'Successfully imported {len(transactions)} transactions',
    #             'count': len(transactions)
    #         })

    #     except requests.RequestException as e:
    #         return Response({'error': f'Failed to fetch data: {str(e)}'}, status=400)
    #     except Exception as e:
    #         return Response({'error': str(e)}, status=400)
        
    ####################### gmo
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        from datetime import datetime
        import csv
        from io import StringIO
        import requests
        import re

        csv_data = request.data.get('csv_data', '')
        sheet_url = request.data.get('sheet_url', '')

        if not csv_data and not sheet_url:
            return Response({'error': 'Either csv_data or sheet_url is required'}, status=400)

        try:
            # ==============================
            # Fetch CSV from Google Sheets
            # ==============================
            if sheet_url:
                if 'pub?output=csv' in sheet_url:
                    csv_url = sheet_url
                else:
                    sheet_id = None
                    gid = '0'

                    if '/d/' in sheet_url:
                        sheet_id = sheet_url.split('/d/')[1].split('/')[0]
                        if 'gid=' in sheet_url:
                            gid = sheet_url.split('gid=')[1].split('&')[0].split('#')[0]

                    if not sheet_id:
                        return Response({'error': 'Invalid Google Sheets URL format'}, status=400)

                    csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'

                response = requests.get(csv_url, timeout=30)
                response.raise_for_status()
                response.encoding = 'utf-8'
                csv_data = response.text

            csv_file = StringIO(csv_data)
            reader = csv.reader(csv_file)

            # skip header
            next(reader, None)

            transactions = []

            for row in reader:
                # Expecting at least 6 columns
                if len(row) < 6:
                    continue

                # ==============================
                # DATE
                # ==============================
                date_str = row[0].strip()
                date_obj = None

                for fmt in (
                    '%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y',
                    '%m/%d/%Y', '%Y%m%d'
                ):
                    try:
                        date_obj = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue

                if not date_obj:
                    continue

                # ==============================
                # WITHDRAW
                # ==============================
                withdraw = 0
                if row[3].strip() and row[3].strip() != '-':
                    try:
                        withdraw = float(row[3].strip())
                    except:
                        pass

                # ==============================
                # DEPOSIT
                # ==============================
                deposit = 0
                if row[2].strip() and row[2].strip() != '-':
                    try:
                        deposit = float(row[2].strip())
                    except:
                        pass

                # ==============================
                # DESCRIPTION (お取り扱い内容)
                # ==============================
                # transaction_id = row[3].strip()

                # # normalize full-width spaces
                # transaction_id = transaction_id.replace('\u3000', ' ')

                # # extract codes like V495093
                # match = re.match(r'^([A-Z]\d+)', transaction_id)

                # if match:
                #     transaction_id = match.group(1)
                # else:
                #     transaction_id = transaction_id

                # transaction_id = transaction_id[:500]

                # ==============================
                # BALANCE
                # ==============================
                balance = 0
                if row[4].strip():
                    try:
                        balance = float(row[4].strip())
                    except:
                        pass
                company_account = None
                if len(row) > 5 and row[5].strip():
                    try:
                        account_id = int(row[5].strip())
                        company_account = CompanyAccount.objects.filter(id=account_id).first()
                    except ValueError:
                        company_account = None

                # ==============================
                # NOTES (メモ)
                # ==============================
                notes = ''
                # notes = row[5].strip() if len(row) > 5 else ''

                transactions.append(Transaction(
                    user=request.user,
                    date=date_obj,
                    withdraw=withdraw,
                    deposit=deposit,
                    balance=balance,
                    description='',
                    transaction_id='',
                    notes=notes,
                    company_account=company_account,
                ))

            Transaction.objects.bulk_create(transactions)

            return Response({
                'message': f'Successfully imported {len(transactions)} transactions',
                'count': len(transactions)
            })

        except requests.RequestException as e:
            return Response({'error': f'Failed to fetch data: {str(e)}'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
