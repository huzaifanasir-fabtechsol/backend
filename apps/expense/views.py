from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction as db_transaction
from django.db.models import Q
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from apps.expense.models import Expense, ExpenseCategory, Restaurant, SparePart
from apps.expense.serializers import ExpenseSerializer, ExpenseCategorySerializer, RestaurantSerializer, SparePartSerializer
from apps.revenue.models import CompanyAccount, Transaction
from apps.revenue.serializers import TransactionSerializer
from apps.account.models import User

class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ExpenseCategory.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)



    @action(detail=False, methods=['get'])
    def all(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Expense.objects.filter(user=self.request.user)

        category = self.request.query_params.get('category')
        date = self.request.query_params.get('date')
        search = self.request.query_params.get('search')

        if category:
            queryset = queryset.filter(category_id=category)
        if date:
            queryset = queryset.filter(date=date)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(category__name__icontains=search) |
                Q(restaurant__name__icontains=search) |
                Q(spare_part__name__icontains=search) |
                Q(spare_part__address__icontains=search)
            )

        return queryset.order_by('-date', '-id')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


    def _parse_excel_date(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        value_str = str(value).strip()
        if not value_str:
            return None

        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d'):
            try:
                return datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_excel_amount(self, value):
        if value is None:
            return None
        value_str = str(value).strip()
        if not value_str:
            return None

        cleaned = value_str.replace(',', '').replace('\u00a5', '').replace('$', '')
        cleaned = cleaned.replace('(', '-').replace(')', '')
        try:
            amount = Decimal(cleaned)
        except (InvalidOperation, TypeError):
            return None

        if amount < 0:
            amount = -amount
        return amount

    @action(detail=False, methods=['get'])
    def search_titles(self, request):
        query = request.query_params.get('q', '').strip()
        if not query:
            return Response([])
        
        titles = Expense.objects.filter(
            user=request.user,
            title__icontains=query
        ).values_list('title', flat=True).distinct()[:10]
        
        return Response(list(titles))

    def _add_watermark(self, canvas, doc, text="INVOICE"):
            canvas.saveState()
            canvas.setFont("HeiseiMin-W3", 80)
            canvas.setFillColor(colors.lightgrey)
            canvas.setFillAlpha(0.15)

            width, height = doc.pagesize
            canvas.translate(width / 2, height / 2)
            canvas.rotate(45)
            canvas.drawCentredString(0, 0, text)

            canvas.restoreState()

    def _add_page_decorations(self, canvas, doc, include_logo=True):
        from django.conf import settings
        import os
        from reportlab.lib.utils import ImageReader
        self._add_watermark(canvas, doc, "Ilyas Sons 合同会社")

        if include_logo:
            logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")
            print("MEDIA_ROOT:", settings.MEDIA_ROOT)
            print("Logo path:", logo_path)
            print("Exists:", os.path.exists(logo_path))
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)

                logo_width = 120
                logo_height = 40

                x = doc.leftMargin
                y = doc.pagesize[1] - logo_height - 10

                canvas.drawImage(
                    logo,
                    x,
                    y,
                    width=logo_width,
                    height=logo_height,
                    mask='auto'
                )

    def _add_first_page_decorations(self, canvas, doc):
        self._add_page_decorations(canvas, doc, include_logo=True)

    def _add_later_page_decorations(self, canvas, doc):
        self._add_page_decorations(canvas, doc, include_logo=False)


    @action(detail=True, methods=['get'])
    def generate_receipt(self, request, pk=None):
        expense = self.get_object()
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        user = request.user

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="expense_{expense.id}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=24, alignment=TA_CENTER, fontName='HeiseiMin-W3')
        elements.append(Paragraph("経費領収書", title_style))
        elements.append(Spacer(1, 15))

        # Company info on right
        header_data = [
            ["", "", user.company_name],
            ["", "", user.company_address],
            ["", "", f"TEL/FAX: {user.company_phone}"],
            ["", "", user.business_registration],
            ["", "", f"Date: {expense.date}"]
        ]
        header_table = Table(header_data, colWidths=[doc.width * 0.4, doc.width * 0.2, doc.width * 0.4])
        header_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (2, 0), (2, -1), 'HeiseiMin-W3')
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 10))

        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.black))
        elements.append(Spacer(1, 15))

        # Expense details
        detail_data = [
            ["タイトル:", expense.title],
            ["額:", f"¥ {expense.amount:,.0f}"],
            ["カテゴリ:", expense.category.name if expense.category else 'N/A'],
            ["説明:", expense.description or '-']
        ]

        if expense.transaction:
            detail_data.extend([
                ["トランザクションID:", str(expense.transaction.transaction_id)],
                ["取引:", expense.transaction.description],
                ["取引金額:", f"¥ {expense.transaction.withdraw:,.0f}"]
            ])

        if expense.restaurant:
            detail_data.append(["レストラン:", expense.restaurant.name])
        if expense.spare_part:
            spare_part_text = expense.spare_part.name
            if expense.spare_part.address:
                spare_part_text = f"{spare_part_text} - {expense.spare_part.address}"
            detail_data.append(["店:", spare_part_text])

        detail_table = Table(detail_data, colWidths=[doc.width * 0.3, doc.width * 0.7])
        detail_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('FONTNAME', (0, 0), (-1, -1), 'HeiseiMin-W3'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
        ]))
        elements.append(detail_table)
        doc.build(
            elements,
            onFirstPage=self._add_first_page_decorations,
            onLaterPages=self._add_later_page_decorations,
        )
        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        user = request.user

        # Apply filters
        queryset = self.get_queryset()
        date = request.query_params.get('date')
        category = request.query_params.get('category')
        search = request.query_params.get('search')

        if date:
            queryset = queryset.filter(date=date)
        if category:
            queryset = queryset.filter(category_id=category)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(category__name__icontains=search) |
                Q(spare_part__name__icontains=search) |
                Q(spare_part__address__icontains=search)
            )

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="expenses.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('タイトル', parent=styles['Normal'], fontSize=24, alignment=TA_CENTER, fontName='HeiseiMin-W3')
        elements.append(Paragraph("経費請求書", title_style))
        elements.append(Spacer(1, 15))

        # Header with company info and filters
        header_data = []
        left_col = []
        if date:
            left_col.append(f"日付: {date}")
        if category:
            cat = ExpenseCategory.objects.filter(id=category).first()
            if cat:
                left_col.append(f"カテゴリ: {cat.name}")
        # left_col.append(f"Generated: {request.user.email}")

        right_col = [user.company_name, user.company_address, f"TEL/FAX: {user.company_phone}", user.business_registration]

        max_rows = max(len(left_col), len(right_col))
        for i in range(max_rows):
            header_data.append([
                left_col[i] if i < len(left_col) else "",
                "",
                right_col[i] if i < len(right_col) else ""
            ])

        header_table = Table(header_data, colWidths=[doc.width * 0.4, doc.width * 0.2, doc.width * 0.4])
        header_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'HeiseiMin-W3')
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 10))

        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.black))
        elements.append(Spacer(1, 15))

        # Table
        cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='HeiseiMin-W3',
            fontSize=8,
            leading=10,
            wordWrap='CJK',
        )
        right_cell_style = ParagraphStyle(
            'TableCellRight',
            parent=cell_style,
            alignment=TA_RIGHT,
        )

        table_data = [['Sr', '日付', 'タイトル', 'カテゴリ', '取引', 'レストラン', 'ショップ', '額']]
        total = 0
        for idx, expense in enumerate(queryset, 1):
            spare_part_text = '-'
            if expense.spare_part:
                spare_part_text = expense.spare_part.name
                if expense.spare_part.address:
                    spare_part_text = f"{spare_part_text} - {expense.spare_part.address}"
            table_data.append([
                str(idx),
                str(expense.date),
                Paragraph(expense.title, cell_style),
                Paragraph(expense.category.name if expense.category else '-', cell_style),
                Paragraph(f"{expense.transaction.transaction_id}" if expense.transaction else '-', cell_style),
                Paragraph(expense.restaurant.name if expense.restaurant else '-', cell_style),
                Paragraph(spare_part_text, cell_style),
                Paragraph(f"¥ {expense.amount:,.0f}", right_cell_style)
            ])
            total += expense.amount

        table_data.append([
            '',
            '',
            '',
            '',
            '',
            '',
            Paragraph('合計:', cell_style),
            Paragraph(f"¥ {total:,.0f}", right_cell_style)
        ])

        table = Table(table_data, colWidths=[25, 55, 90, 70, 65, 95, 85, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'HeiseiMin-W3'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'HeiseiMin-W3'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        elements.append(table)

        doc.build(
            elements,
            onFirstPage=self._add_first_page_decorations,
            onLaterPages=self._add_later_page_decorations,
        )
        return response

    @action(detail=False, methods=['get'])
    def available_transactions(self, request):
        search = request.query_params.get('search', '')
        date = request.query_params.get('date', '')
        account_id = request.query_params.get('account_id', '')

        queryset = Transaction.objects.filter(user=request.user)

        if account_id:
            queryset = queryset.filter(company_account_id=account_id)
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(notes__icontains=search) |
                Q(balance__icontains=search) |
                Q(deposit__icontains=search) |
                Q(withdraw__icontains=search)
            )
        if date:
            queryset = queryset.filter(date=date)

        serializer = TransactionSerializer(queryset, many=True)
        return Response(serializer.data)


    @action(detail=False, methods=['post'], url_path='bulk-import-xls-expenses')
    def bulk_import_xls_expenses(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        category = ExpenseCategory.objects.filter(id=17, user=request.user).first()
        if not category:
            return Response({'error': 'Expense category 17 not found for this user'}, status=status.HTTP_400_BAD_REQUEST)

        account_id = request.data.get('company_account_id')
        company_account = None
        if account_id:
            company_account = CompanyAccount.objects.filter(id=account_id, user=request.user).first()
            if not company_account:
                return Response({'error': 'company_account_id is invalid'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception as exc:
            return Response(
                {'error': f'Unable to read Excel file. Upload a valid .xlsx file. {str(exc)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            return Response({'error': 'Excel file has no data rows'}, status=status.HTTP_400_BAD_REQUEST)

        default_account = company_account or CompanyAccount.objects.filter(user=request.user).order_by('id').first()
        missing_account_needed = False
        created_transactions = 0
        reused_transactions = 0
        created_expenses = 0
        skipped_rows = 0
        row_errors = []

        with db_transaction.atomic():
            for row_index, row in enumerate(rows[1:], start=2):
                row_values = list(row or [])
                if len(row_values) < 3:
                    skipped_rows += 1
                    row_errors.append(f'Row {row_index}: expected at least 3 columns')
                    continue

                tx_id_raw = row_values[-1]
                tx_id = str(tx_id_raw).strip() if tx_id_raw is not None else ''
                tx_id = tx_id[:500]
                if not tx_id:
                    skipped_rows += 1
                    row_errors.append(f'Row {row_index}: transaction id is empty')
                    continue

                parsed_date = self._parse_excel_date(row_values[0])
                if not parsed_date:
                    skipped_rows += 1
                    row_errors.append(f'Row {row_index}: invalid date "{row_values[0]}"')
                    continue

                amount = self._parse_excel_amount(row_values[1])
                if amount is None:
                    skipped_rows += 1
                    row_errors.append(f'Row {row_index}: invalid amount "{row_values[1]}"')
                    continue

                existing_tx = Transaction.objects.filter(
                    user=request.user,
                    transaction_id=tx_id
                ).order_by('id').first()

                if existing_tx:
                    tx = existing_tx
                    reused_transactions += 1
                else:
                    if not default_account:
                        missing_account_needed = True
                        skipped_rows += 1
                        row_errors.append(
                            f'Row {row_index}: transaction "{tx_id}" not found and no company account available to create it'
                        )
                        continue

                    tx = Transaction.objects.create(
                        user=request.user,
                        date=parsed_date,
                        transaction_id=tx_id,
                        withdraw=amount,
                        deposit=Decimal('0'),
                        balance=Decimal('0'),
                        description='Imported from expense xls',
                        notes='',
                        company_account=default_account,
                    )
                    created_transactions += 1

                Expense.objects.create(
                    user=request.user,
                    title='Highway',
                    category=category,
                    date=parsed_date,
                    amount=amount,
                    transaction=tx,
                    description='',
                )
                created_expenses += 1

        response_data = {
            'message': 'Bulk import completed',
            'created_transactions': created_transactions,
            'reused_transactions': reused_transactions,
            'created_expenses': created_expenses,
            'skipped_rows': skipped_rows,
            'errors': row_errors[:100],
        }
        if missing_account_needed:
            response_data['note'] = (
                'Provide company_account_id in the request or create a company account to allow creating new transactions.'
            )
        return Response(response_data, status=status.HTTP_200_OK)

class RestaurantViewSet(viewsets.ModelViewSet):
    serializer_class = RestaurantSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Restaurant.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(location__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class SparePartViewSet(viewsets.ModelViewSet):
    serializer_class = SparePartSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = SparePart.objects.filter(user=self.request.user)
        search = self.request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(address__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
